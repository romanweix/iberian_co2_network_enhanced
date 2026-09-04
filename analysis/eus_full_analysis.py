# eus_full_analysis.py
#
# Full EUS (base-case utilization scenario) analysis across every
# stochastic_results_*.xlsx run in analysis/results_data/. Builds on
# compare_scenarios.py (same KPI summary / cost breakdown comparison
# columns and logic) and adds a few extra cross-scenario comparisons:
# a condensed cost-category view (overall, and vs. each phase's benchmark
# run), the network build-out timeline, an installed-diameter distribution
# pivot, cost-efficiency metrics (EUR per km built, EUR per tonne of CO2
# stored), why boosters were needed (pressure vs. temperature), the total
# pipeline volume (pipe itself, no insulation) implied by each scenario's
# installed pipes, and the insulation shell volume implied by each
# scenario's insulated pipes.
#
# Usage (from the repo root):
#   python -m analysis.eus_full_analysis

import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from analysis.compare_scenarios import (
    compare_scenarios,
    _parse_scenario_label,
    _load_cost_breakdown,
    _gross_net_cost,
    MEMO_ONLY_ROWS,
    GROSS_COST_EXCLUDED_ROWS,
    M_EUR_TO_BN_EUR_SCALE,
)

RESULTS_DIR = Path("analysis/results_data")
UTIL = "EUS"
OUT_PATH = Path("analysis/EUS_full_analysis.xlsx")

# Benchmark runs outside the insulation-surcharge sweep: not "supercritical"
# / "liquid" + surcharge like the rest, so label them explicitly instead of
# letting _parse_scenario_label's default (supercritical, no surcharge)
# quietly conflate them with "noins".
BENCHMARK_PHASE_LABELS = {
    "bm_sco2": "benchmark (sCO2 network)",
    "bm_dense_2": "benchmark (dense-phase network)",
}

# Supercritical-phase sweep scenarios and their benchmark; liquid-phase
# sweep scenarios and their benchmark (dense-phase network run).
SUPERCRITICAL_SCENARIOS = ["noins", "ins20", "ins40", "ins60", "ins80", "ins100", "ins150"]
SUPERCRITICAL_BENCHMARK = "bm_sco2"
LIQUID_SCENARIOS = ["liq_noins", "liq_ins20", "liq_ins40", "liq_ins60"]
LIQUID_BENCHMARK = "bm_dense_2"

# Model constants from iberian_co2_network/data.py's scenario data dict
# (same for every run): minimum allowed pressure/temperature anywhere in
# the network. theta_min differs by phase (SCO2_PHASE flag at the top of
# data.py): 32 degC for the supercritical runs, 15 degC for the liquid runs.
P_MIN_BAR = 100.0
THETA_MIN_SUPERCRITICAL_C = 32.0
THETA_MIN_LIQUID_C = 15.0

# Pipe insulation shell thickness, per user specification (not itself a
# modeled quantity -- the model tracks only insulated-or-not per pipe).
INSULATION_THICKNESS_M = 0.15
INCH_TO_M = 0.0254

# Shared with plot_booster_vs_insulation.py, which reuses build_booster_reasons().
REASON_ORDER = ["Pressure only", "Temperature only", "Both", "Inconclusive (installation-year snapshot)"]

# High-level buckets for the condensed "Cost category breakdown" tab.
# Every raw Concept row from {util} - Cost breakdown must appear in exactly
# one bucket (checked in build_cost_category_breakdown), so the bucket
# totals reconcile with the Gross/Net cost figures in the KPI summary.
COST_CATEGORY_MAP = {
    "Capture cost": "Capture cost",
    "Penalty for uncaptured emissions": "Penalty for uncaptured emissions",
    "Injection cost": "Injection",
    "CAPEX onshore pipe": "Pipeline CAPEX",
    "CAPEX offshore pipe": "Pipeline CAPEX",
    "OPEX onshore pipe": "Pipeline OPEX",
    "OPEX offshore pipe": "Pipeline OPEX",
    "CAPEX onshore pipe (insulation only)": "Insulation CAPEX (memo)",
    "OPEX onshore pipe (insulation only)": "Insulation OPEX (memo)",
    "CAPEX initial boosting stations": "Boosting CAPEX",
    "CAPEX additional boosting stations": "Boosting CAPEX",
    "OPEX initial boosting stations": "Boosting OPEX",
    "OPEX additional boosting stations": "Boosting OPEX",
    "Electricity cost initial boosting stations": "Boosting electricity",
    "Electricity cost additional boosting stations": "Boosting electricity",
    "CO2 sales revenue": "CO2 sales revenue",
    # Only present in the older benchmark runs (bm_sco2, bm_dense_2), which
    # predate the pipeline-insulation feature; always 0 in this data set.
    # Counted towards Gross/Net cost like the other real (non-memo) rows,
    # since compare_scenarios._gross_net_cost() doesn't exclude it either.
    "Shipping cost": "Shipping",
}

# Order to display bucket rows in; the memo rows are informational only
# (already included in Pipeline CAPEX/OPEX) and excluded from the Gross/Net
# subtotal, matching MEMO_ONLY_ROWS in compare_scenarios.py.
CATEGORY_ORDER = [
    "Injection",
    "Shipping",
    "Pipeline CAPEX",
    "Pipeline OPEX",
    "Insulation CAPEX (memo)",
    "Insulation OPEX (memo)",
    "Boosting CAPEX",
    "Boosting OPEX",
    "Boosting electricity",
    "Gross cost (Bn€)",
    "CO2 sales revenue",
    "Net cost (Bn€)",
    "Capture cost (excluded from Gross/Net)",
    "Penalty for uncaptured emissions (excluded from Gross/Net)",
]


def scenario_files():
    return sorted(RESULTS_DIR.glob("*_stochastic_results_theta_*.xlsx"))


def labeled_scenarios():
    """(file, scenario_key, phase_label) for every results file, with
    benchmark runs given an explicit phase label instead of the sweep's
    default."""
    out = []
    for f in scenario_files():
        scenario_key, phase, surcharge_pct = _parse_scenario_label(f.name)
        phase_label = BENCHMARK_PHASE_LABELS.get(scenario_key, phase)
        out.append((f, scenario_key, phase_label, surcharge_pct))
    return out


def build_cost_category_breakdown() -> pd.DataFrame:
    rows = {}
    for f, scenario_key, phase_label, surcharge_pct in labeled_scenarios():
        cost_series = _load_cost_breakdown(f, UTIL)

        unmapped = [c for c in cost_series.index if c not in COST_CATEGORY_MAP]
        if unmapped:
            raise ValueError(f"{f.name}: unmapped cost concept(s) {unmapped}; "
                              "update COST_CATEGORY_MAP in eus_full_analysis.py")

        bucket_totals = {cat: 0.0 for cat in COST_CATEGORY_MAP.values()}
        for concept, value in cost_series.items():
            bucket_totals[COST_CATEGORY_MAP[concept]] += value

        gross, net = _gross_net_cost(cost_series)

        col = {
            "Injection": bucket_totals["Injection"] / M_EUR_TO_BN_EUR_SCALE,
            "Shipping": bucket_totals["Shipping"] / M_EUR_TO_BN_EUR_SCALE,
            "Pipeline CAPEX": bucket_totals["Pipeline CAPEX"] / M_EUR_TO_BN_EUR_SCALE,
            "Pipeline OPEX": bucket_totals["Pipeline OPEX"] / M_EUR_TO_BN_EUR_SCALE,
            "Insulation CAPEX (memo)": bucket_totals["Insulation CAPEX (memo)"] / M_EUR_TO_BN_EUR_SCALE,
            "Insulation OPEX (memo)": bucket_totals["Insulation OPEX (memo)"] / M_EUR_TO_BN_EUR_SCALE,
            "Boosting CAPEX": bucket_totals["Boosting CAPEX"] / M_EUR_TO_BN_EUR_SCALE,
            "Boosting OPEX": bucket_totals["Boosting OPEX"] / M_EUR_TO_BN_EUR_SCALE,
            "Boosting electricity": bucket_totals["Boosting electricity"] / M_EUR_TO_BN_EUR_SCALE,
            "Gross cost (Bn€)": gross,
            "CO2 sales revenue": bucket_totals["CO2 sales revenue"] / M_EUR_TO_BN_EUR_SCALE,
            "Net cost (Bn€)": net,
            "Capture cost (excluded from Gross/Net)": bucket_totals["Capture cost"] / M_EUR_TO_BN_EUR_SCALE,
            "Penalty for uncaptured emissions (excluded from Gross/Net)":
                bucket_totals["Penalty for uncaptured emissions"] / M_EUR_TO_BN_EUR_SCALE,
        }
        rows[scenario_key] = col

    df = pd.DataFrame(rows).reindex(CATEGORY_ORDER)
    df.index.name = "Cost category [Bn€]"
    return df


def build_network_length_by_year() -> pd.DataFrame:
    """EUS row of the 'Network length per scenario' sheet (cumulative
    installed pipe length by year, km) for every results file."""
    rows = {}
    for f, scenario_key, phase_label, surcharge_pct in labeled_scenarios():
        df = pd.read_excel(f, sheet_name="Network length per scenario")
        eus_row = df[df["Scenario"] == UTIL].iloc[0]
        rows[scenario_key] = eus_row.drop(labels=["Scenario"])
    out = pd.DataFrame(rows).T
    out.index.name = "Scenario"
    out.columns.name = None
    return out


def build_diameter_distribution() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Pivot of installed pipe count and installed length [km] by diameter
    [inch] (rows) x scenario (columns), from the {util} - Pipes sheet."""
    count_rows = {}
    length_rows = {}
    for f, scenario_key, phase_label, surcharge_pct in labeled_scenarios():
        df = pd.read_excel(f, sheet_name=f"{UTIL} - Pipes")
        installed = df[df["Installed"] == 1]
        by_diam = installed.groupby("Diameter [inch]")
        count_rows[scenario_key] = by_diam.size()
        length_rows[scenario_key] = by_diam["Longitude [km]"].sum()

    count_df = pd.DataFrame(count_rows).fillna(0)
    length_df = pd.DataFrame(length_rows).fillna(0)
    count_df.index = [f'{int(d)}"' for d in count_df.index]
    length_df.index = [f'{int(d)}"' for d in length_df.index]
    count_df.index.name = "Diameter"
    length_df.index.name = "Diameter"

    def _sort(df):
        df = df.copy()
        df["_sort"] = [int(i.rstrip('"')) for i in df.index]
        return df.sort_values("_sort").drop(columns="_sort")

    return _sort(count_df), _sort(length_df)


def build_cost_efficiency(kpi_summary: pd.DataFrame) -> pd.DataFrame:
    """Net cost per km built and per tonne of CO2 stored, using the
    {util} - Sinks capacity evolution sheet's Initial vs. Final capacity
    (Mt) as the cumulative amount of CO2 injected over the horizon."""
    stored_mt = {}
    for f, scenario_key, phase_label, surcharge_pct in labeled_scenarios():
        df = pd.read_excel(f, sheet_name=f"{UTIL} - Sinks capacity evolution")
        stored_mt[scenario_key] = (df["Initial capacity [Mt]"] - df["Final capacity [Mt]"]).sum()

    rows = []
    for _, r in kpi_summary.iterrows():
        scenario_key = r["Scenario"]
        mt = stored_mt.get(scenario_key, 0.0)
        net_cost_bn = r["Net cost [Bn€]"]
        length_km = r["Total pipe length [km]"]
        rows.append({
            "Scenario": scenario_key,
            "Phase": r["Phase"],
            "Insulation surcharge [%]": r["Insulation surcharge [%]"],
            "Net cost [Bn€]": net_cost_bn,
            "Total pipe length [km]": length_km,
            "Total CO2 stored [Mt]": mt,
            "Net cost per km [k€/km]": (net_cost_bn * 1e6 / length_km) if length_km else None,
            "Net cost per tCO2 stored [€/t]": (net_cost_bn * 1000.0 / mt) if mt else None,
        })
    return pd.DataFrame(rows)


def build_cost_category_vs_benchmark(cost_category: pd.DataFrame, scenarios: list[str],
                                      benchmark: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Cost-category values for one phase's sweep scenarios plus its
    benchmark run, and each scenario's percent deviation from that
    benchmark per cost category ((scenario - benchmark) / benchmark * 100).
    A benchmark value of ~0 (e.g. insulation rows for the pre-insulation
    benchmark runs) makes the deviation undefined, left blank rather than
    a divide-by-zero infinity."""
    cols = [benchmark] + scenarios
    values = cost_category[cols]

    pct = pd.DataFrame(index=values.index, columns=cols, dtype=float)
    bench_col = values[benchmark]
    for col in cols:
        with np.errstate(divide="ignore", invalid="ignore"):
            dev = np.where(bench_col.abs() > 1e-9, (values[col] - bench_col) / bench_col * 100.0, np.nan)
        pct[col] = dev
    pct.index.name = values.index.name
    return values, pct


def build_booster_reasons() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    """For every installed pipe with at least one booster, classify why it
    was needed: developed_solution.py reports "Pressure/Temperature at
    [critical point]" and "Lowest pressure/temperature" using the RAW
    simulated pressure/temperature drop from the pipe's own origin value at
    its installation year -- i.e. what that point would be WITHOUT the
    booster's compensation (see the Pi_pmax/Pi_dest/Theta_pmax/Theta_cold
    usage in create_pipe_summary()). So a reported value already below
    p_min/theta_min directly signals that dimension was (at least at the
    pipe's own installation year) the reason a booster was required.

    Caveat, noted in the output: the booster's own build year can be later
    than the pipe's installation year (flows, and so required pressure,
    change over time), so a handful of boosted pipes can show neither
    value below threshold at the installation-year snapshot used here --
    labeled "Inconclusive" rather than mis-assigned.

    Returns (detail_df, count_matrix, booster_count_matrix, excluded_scenarios).
    """
    detail_rows = []
    excluded = []

    for f, scenario_key, phase_label, surcharge_pct in labeled_scenarios():
        df = pd.read_excel(f, sheet_name=f"{UTIL} - Pipes")
        if "Temperature at critical pressure point [°C]" not in df.columns:
            excluded.append(scenario_key)
            continue

        theta_min = THETA_MIN_LIQUID_C if scenario_key.startswith("liq_") else THETA_MIN_SUPERCRITICAL_C
        boosted = df[df["Number of boosters"].fillna(0) > 0]

        for _, r in boosted.iterrows():
            p_flag = (r["Pressure at highest point [bar]"] < P_MIN_BAR) or (r["Lowest pressure [bar]"] < P_MIN_BAR)
            t_flag = (r["Temperature at critical pressure point [°C]"] < theta_min) or \
                     (r["Lowest temperature [°C]"] < theta_min)
            if p_flag and t_flag:
                reason = "Both"
            elif p_flag:
                reason = "Pressure only"
            elif t_flag:
                reason = "Temperature only"
            else:
                reason = "Inconclusive (installation-year snapshot)"

            detail_rows.append({
                "Scenario": scenario_key,
                "Phase": phase_label,
                "Pipe ID": r["Pipe ID"],
                "Number of boosters": r["Number of boosters"],
                "Pressure at highest point [bar]": r["Pressure at highest point [bar]"],
                "Lowest pressure [bar]": r["Lowest pressure [bar]"],
                "p_min [bar]": P_MIN_BAR,
                "Temperature at critical pressure point [°C]": r["Temperature at critical pressure point [°C]"],
                "Lowest temperature [°C]": r["Lowest temperature [°C]"],
                "theta_min [°C]": theta_min,
                "Reason": reason,
            })

    detail_df = pd.DataFrame(detail_rows)

    valid_scenarios = [sk for _, sk, _, _ in labeled_scenarios() if sk not in excluded]
    count_matrix = pd.DataFrame(0, index=REASON_ORDER, columns=valid_scenarios)
    boostercount_matrix = pd.DataFrame(0.0, index=REASON_ORDER, columns=valid_scenarios)
    if not detail_df.empty:
        pivot_n = detail_df.pivot_table(index="Reason", columns="Scenario", values="Pipe ID",
                                         aggfunc="count", fill_value=0)
        pivot_b = detail_df.pivot_table(index="Reason", columns="Scenario", values="Number of boosters",
                                         aggfunc="sum", fill_value=0)
        count_matrix.update(pivot_n)
        boostercount_matrix.update(pivot_b)
    count_matrix.index.name = "Reason for booster"
    boostercount_matrix.index.name = "Reason for booster"

    return detail_df, count_matrix, boostercount_matrix, excluded


def _insulation_shell_volume_m3(diameter_inch: pd.Series, length_km: pd.Series) -> pd.Series:
    """Annular shell volume of a 15 cm insulation layer wrapped around each
    pipe: V = pi * ((r + t)^2 - r^2) * L, with r the pipe's outer radius
    (from its nominal diameter -- no wall-thickness data is modeled) and L
    its installed length."""
    r = (diameter_inch * INCH_TO_M) / 2.0
    t = INSULATION_THICKNESS_M
    length_m = length_km * 1000.0
    return math.pi * ((r + t) ** 2 - r ** 2) * length_m


def _pipe_volume_m3(diameter_inch: pd.Series, length_km: pd.Series) -> pd.Series:
    """Enclosed volume of the pipe itself, without any insulation shell:
    V = pi * r^2 * L, with r the pipe's nominal radius (from its nominal
    diameter -- no wall-thickness data is modeled) and L its installed
    length."""
    r = (diameter_inch * INCH_TO_M) / 2.0
    length_m = length_km * 1000.0
    return math.pi * (r ** 2) * length_m


def build_pipe_volume() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Per-scenario pipeline volume (m^3): the pipe's own enclosed volume
    from its nominal diameter and installed length, independent of any
    insulation shell, plus a diameter x scenario breakdown. Unlike
    build_insulation_volume(), this covers every installed pipe in every
    scenario -- diameter and length are always exported, even for the
    pre-insulation-feature benchmark runs."""
    summary_rows = []
    volume_by_diam = {}

    for f, scenario_key, phase_label, surcharge_pct in labeled_scenarios():
        df = pd.read_excel(f, sheet_name=f"{UTIL} - Pipes")
        installed = df[df["Installed"] == 1].copy()
        installed["_volume_m3"] = _pipe_volume_m3(installed["Diameter [inch]"], installed["Longitude [km]"])

        summary_rows.append({
            "Scenario": scenario_key,
            "Phase": phase_label,
            "Installed pipes": len(installed),
            "Total pipe length [km]": installed["Longitude [km]"].sum(),
            "Total pipe volume [m³]": installed["_volume_m3"].sum(),
        })

        volume_by_diam[scenario_key] = installed.groupby("Diameter [inch]")["_volume_m3"].sum()

    summary_df = pd.DataFrame(summary_rows)

    diam_matrix = pd.DataFrame(volume_by_diam).fillna(0.0)
    diam_matrix.index = [f'{int(d)}"' for d in diam_matrix.index]
    diam_matrix.index.name = "Diameter"
    diam_matrix["_sort"] = [int(i.rstrip('"')) for i in diam_matrix.index]
    diam_matrix = diam_matrix.sort_values("_sort").drop(columns="_sort")

    return summary_df, diam_matrix


def build_insulation_volume() -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Per-scenario insulation shell volume (m^3), plus a diameter x
    scenario breakdown for the scenarios where insulation is modeled at
    all (the two pre-insulation-feature benchmark runs are reported
    separately, not silently as zero)."""
    summary_rows = []
    excluded = []
    volume_by_diam = {}

    for f, scenario_key, phase_label, surcharge_pct in labeled_scenarios():
        df = pd.read_excel(f, sheet_name=f"{UTIL} - Pipes")
        if "Insulated" not in df.columns:
            excluded.append(scenario_key)
            summary_rows.append({
                "Scenario": scenario_key,
                "Phase": phase_label,
                "Insulated pipes": None,
                "Insulated length [km]": None,
                "Insulation volume [m³]": None,
                "Note": "n/a -- insulation not modeled in this run (pre-insulation-feature benchmark)",
            })
            continue

        installed = df[df["Installed"] == 1]
        insulated = installed[installed["Insulated"] == True].copy()  # noqa: E712
        insulated["_volume_m3"] = _insulation_shell_volume_m3(
            insulated["Diameter [inch]"], insulated["Longitude [km]"])

        summary_rows.append({
            "Scenario": scenario_key,
            "Phase": phase_label,
            "Insulated pipes": len(insulated),
            "Insulated length [km]": insulated["Longitude [km]"].sum(),
            "Insulation volume [m³]": insulated["_volume_m3"].sum(),
            "Note": "",
        })

        volume_by_diam[scenario_key] = insulated.groupby("Diameter [inch]")["_volume_m3"].sum()

    summary_df = pd.DataFrame(summary_rows)

    diam_matrix = pd.DataFrame(volume_by_diam).fillna(0.0)
    diam_matrix.index = [f'{int(d)}"' for d in diam_matrix.index]
    diam_matrix.index.name = "Diameter"
    diam_matrix["_sort"] = [int(i.rstrip('"')) for i in diam_matrix.index]
    diam_matrix = diam_matrix.sort_values("_sort").drop(columns="_sort")

    return summary_df, diam_matrix, excluded


def build_booster_vs_insulation(kpi_summary: pd.DataFrame, booster_boostcount: pd.DataFrame,
                                 insulation_summary: pd.DataFrame) -> pd.DataFrame:
    """Per-scenario juxtaposition of the two thermal/pressure-management
    strategies the model can trade off against each other: build a booster
    station, or insulate the pipe so it needs boosting less. One row per
    scenario, combining kpi_summary's Total boosters with the reason
    breakdown from build_booster_reasons() and the volumes from
    build_insulation_volume(); scenarios excluded from either (the two
    pre-insulation-feature benchmark runs) get blanks there, not zeros."""
    insulation_by_scenario = insulation_summary.set_index("Scenario")

    rows = []
    for _, r in kpi_summary.iterrows():
        sk = r["Scenario"]
        row = {
            "Scenario": sk,
            "Phase": r["Phase"],
            "Insulation surcharge [%]": r["Insulation surcharge [%]"],
            "Total boosters": r["Total boosters"],
        }
        if sk in booster_boostcount.columns:
            for reason in REASON_ORDER:
                row[f"Boosters -- {reason}"] = booster_boostcount.loc[reason, sk]
        else:
            for reason in REASON_ORDER:
                row[f"Boosters -- {reason}"] = None

        ins = insulation_by_scenario.loc[sk]
        row["Insulated pipes"] = ins["Insulated pipes"]
        row["Insulated length [km]"] = ins["Insulated length [km]"]
        row["Insulation volume [m³]"] = ins["Insulation volume [m³]"]
        rows.append(row)

    return pd.DataFrame(rows)


def _autosize(ws, min_width=10, max_width=42):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def _style_header(ws, n_cols, header_row=1):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _write_blocks(writer, sheet_name, blocks):
    """Write several (title, DataFrame, index) tables stacked vertically on
    one sheet: a bold title row above each table's header, and a blank row
    between tables. Returns the 1-indexed (title_row, header_row) for each
    block, for _style_blocks to style afterwards."""
    layout = []
    row_cursor = 0  # next free 0-indexed row
    for title, df, index in blocks:
        title_row = row_cursor + 1
        header_startrow = row_cursor + 1
        df.to_excel(writer, sheet_name=sheet_name, startrow=header_startrow, index=index)
        header_row = header_startrow + 1
        layout.append({"title": title, "title_row": title_row, "header_row": header_row})
        row_cursor = header_startrow + 1 + len(df) + 1  # header + data rows + blank gap
    return layout


def _style_blocks(ws, layout, n_cols):
    for b in layout:
        cell = ws.cell(row=b["title_row"], column=1, value=b["title"])
        cell.font = Font(bold=True, italic=True)
        _style_header(ws, n_cols, header_row=b["header_row"])
    ws.freeze_panes = None
    _format_numbers(ws)
    _autosize(ws)


def _percent_format_block(ws, layout_entry, n_rows, n_cols):
    """Override _format_numbers' default on one block's data rows with a
    signed-percent display (values are already in percent, e.g. 12.5 -> +12.50%)."""
    start = layout_entry["header_row"] + 1
    for r in range(start, start + n_rows):
        for c in range(2, n_cols + 1):  # column 1 is the row label
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = '+0.00"%";-0.00"%";0.00"%"'


def _bold_rows_by_label(ws, labels: set, label_col: int = 1):
    """Bold every cell in any row whose first column matches one of
    `labels` (e.g. the Gross/Net cost subtotal rows)."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=label_col).value in labels:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(bold=True, italic=cell.font.italic)


def _format_numbers(ws):
    """Two-decimal, thousands-separated display for non-integer numeric
    cells; thousands-separated whole numbers for integer-valued cells
    (pipe/booster counts). Leaves text and blank cells untouched."""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, bool) or v is None:
                continue
            if isinstance(v, (int, float)):
                is_whole = float(v).is_integer()
                cell.number_format = "#,##0" if is_whole else "#,##0.00"


def main():
    result = compare_scenarios(RESULTS_DIR, UTIL)
    cost_comparison, kpi_summary = result["cost_comparison"], result["kpi_summary"]

    # Apply the explicit benchmark phase labels to the KPI summary as well,
    # for consistency with the other tabs.
    kpi_summary = kpi_summary.copy()
    kpi_summary["Phase"] = kpi_summary.apply(
        lambda r: BENCHMARK_PHASE_LABELS.get(r["Scenario"], r["Phase"]), axis=1)

    cost_category = build_cost_category_breakdown()
    network_length = build_network_length_by_year()
    diam_count, diam_length = build_diameter_distribution()
    cost_efficiency = build_cost_efficiency(kpi_summary)

    sc_values, sc_pct = build_cost_category_vs_benchmark(
        cost_category, SUPERCRITICAL_SCENARIOS, SUPERCRITICAL_BENCHMARK)
    sc_values = sc_values.rename(columns={SUPERCRITICAL_BENCHMARK: f"{SUPERCRITICAL_BENCHMARK} (benchmark)"})
    sc_pct = sc_pct.rename(columns={SUPERCRITICAL_BENCHMARK: f"{SUPERCRITICAL_BENCHMARK} (benchmark)"})

    liq_values, liq_pct = build_cost_category_vs_benchmark(
        cost_category, LIQUID_SCENARIOS, LIQUID_BENCHMARK)
    liq_values = liq_values.rename(columns={LIQUID_BENCHMARK: f"{LIQUID_BENCHMARK} (benchmark)"})
    liq_pct = liq_pct.rename(columns={LIQUID_BENCHMARK: f"{LIQUID_BENCHMARK} (benchmark)"})

    booster_detail, booster_count, booster_boostcount, booster_excluded = build_booster_reasons()
    booster_note = pd.DataFrame({"Note": [
        "Reason is derived from the pipe's own Pressure/Temperature at critical point and "
        "Lowest pressure/temperature columns, which developed_solution.py computes from the raw "
        "simulated drop at the pipe's installation year -- i.e. the value WITHOUT the booster's "
        "compensation. A value already below p_min/theta_min there means that dimension needed "
        "the booster. p_min = 100 bar for every run; theta_min = 32°C (supercritical) / 15°C (liquid).",
        "'Inconclusive': neither value is below threshold at the pipe's installation-year snapshot -- "
        "can happen when the booster's own build year is later than the pipe's installation year "
        "(flows, and so the required pressure/temperature, change over time).",
        f"Excluded (older schema, no pressure/temperature columns exported): {', '.join(booster_excluded) or 'none'}.",
    ]})

    pipe_volume_summary, pipe_volume_by_diam = build_pipe_volume()
    pipe_volume_note = pd.DataFrame({"Note": [
        "Pipeline volume = pi * r^2 * length, r = nominal pipe diameter / 2 (no pipe-wall-thickness "
        "data is modeled, so the nominal diameter is used as the pipe's inner diameter). This is the "
        "pipe's own enclosed volume only -- it excludes the insulation shell (see the 'Insulation "
        "volume' tab) and is reported for every installed pipe in every scenario.",
    ]})

    insulation_summary, insulation_by_diam, insulation_excluded = build_insulation_volume()
    insulation_note = pd.DataFrame({"Note": [
        "Insulation shell volume = pi * ((r + 0.15m)^2 - r^2) * length, r = nominal pipe diameter / 2 "
        "(no pipe-wall-thickness data is modeled, so the nominal diameter is used as the pipe's outer "
        "diameter). 0.15 m insulation thickness per user specification.",
        f"Excluded from both tables (older schema, insulation not modeled): {', '.join(insulation_excluded) or 'none'}.",
    ]})

    booster_vs_insulation = build_booster_vs_insulation(kpi_summary, booster_boostcount, insulation_summary)
    bvi_note = pd.DataFrame({"Note": [
        "Juxtaposes the model's two ways of keeping a pipe above p_min/theta_min: build a booster "
        "station, or insulate the pipe so it loses less heat and needs boosting less. Booster reason "
        "columns and insulation columns are blank for bm_sco2/bm_dense_2 (older schema -- see the "
        "'Booster reason' and 'Insulation volume' tabs' notes); Total boosters is still shown for them.",
    ]})

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        kpi_summary.to_excel(writer, sheet_name="KPI summary", index=False)
        cost_comparison.to_excel(writer, sheet_name="Cost breakdown comparison")
        cost_category.to_excel(writer, sheet_name="Cost category breakdown")
        network_length.to_excel(writer, sheet_name="Network length by year")

        diam_layout = _write_blocks(writer, "Diameter distribution", [
            ("Installed pipe count by diameter", diam_count, True),
            ("Installed pipe length [km] by diameter", diam_length, True),
        ])
        cost_efficiency.to_excel(writer, sheet_name="Cost efficiency", index=False)

        sc_layout = _write_blocks(writer, "Cost category - supercritical", [
            ("Cost category breakdown [Bn€] -- supercritical scenarios, benchmark = bm_sco2", sc_values, True),
            ("% deviation vs. supercritical benchmark (bm_sco2)", sc_pct, True),
        ])
        liq_layout = _write_blocks(writer, "Cost category - liquid", [
            ("Cost category breakdown [Bn€] -- liquid scenarios, benchmark = bm_dense_2", liq_values, True),
            ("% deviation vs. dense-phase benchmark (bm_dense_2)", liq_pct, True),
        ])

        booster_layout = _write_blocks(writer, "Booster reason", [
            ("Per-pipe booster reason detail (EUS, boosted pipes only)", booster_detail, False),
            ("Number of boosted pipes by reason x scenario", booster_count, True),
            ("Total booster count by reason x scenario (sum of 'Number of boosters')", booster_boostcount, True),
            ("Notes", booster_note, False),
        ])

        pipe_volume_layout = _write_blocks(writer, "Pipe volume", [
            ("Pipeline volume by scenario (no insulation)", pipe_volume_summary, False),
            ("Pipeline volume [m³] by diameter x scenario", pipe_volume_by_diam, True),
            ("Notes", pipe_volume_note, False),
        ])

        insulation_layout = _write_blocks(writer, "Insulation volume", [
            ("Insulation shell volume by scenario (15 cm layer)", insulation_summary, False),
            ("Insulation shell volume [m³] by diameter x scenario", insulation_by_diam, True),
            ("Notes", insulation_note, False),
        ])

        bvi_layout = _write_blocks(writer, "Booster vs insulation", [
            ("Boosters vs. insulation by scenario (EUS)", booster_vs_insulation, False),
            ("Notes", bvi_note, False),
        ])

    wb = openpyxl.load_workbook(OUT_PATH)

    ws = wb["KPI summary"]
    _style_header(ws, ws.max_column)
    _format_numbers(ws)
    _autosize(ws)

    ws = wb["Cost breakdown comparison"]
    _style_header(ws, ws.max_column)
    _format_numbers(ws)
    _autosize(ws)

    ws = wb["Cost category breakdown"]
    _style_header(ws, ws.max_column)
    _format_numbers(ws)
    _autosize(ws)
    _bold_rows_by_label(ws, {"Gross cost (Bn€)", "Net cost (Bn€)"})

    ws = wb["Network length by year"]
    _style_header(ws, ws.max_column)
    _format_numbers(ws)
    _autosize(ws)

    ws = wb["Diameter distribution"]
    _style_blocks(ws, diam_layout, ws.max_column)

    ws = wb["Cost efficiency"]
    _style_header(ws, ws.max_column)
    _format_numbers(ws)
    _autosize(ws)

    ws = wb["Cost category - supercritical"]
    _style_blocks(ws, sc_layout, ws.max_column)
    _bold_rows_by_label(ws, {"Gross cost (Bn€)", "Net cost (Bn€)"})
    _percent_format_block(ws, sc_layout[1], len(sc_pct), ws.max_column)

    ws = wb["Cost category - liquid"]
    _style_blocks(ws, liq_layout, ws.max_column)
    _bold_rows_by_label(ws, {"Gross cost (Bn€)", "Net cost (Bn€)"})
    _percent_format_block(ws, liq_layout[1], len(liq_pct), ws.max_column)

    ws = wb["Booster reason"]
    _style_blocks(ws, booster_layout, ws.max_column)

    ws = wb["Pipe volume"]
    _style_blocks(ws, pipe_volume_layout, ws.max_column)

    ws = wb["Insulation volume"]
    _style_blocks(ws, insulation_layout, ws.max_column)

    ws = wb["Booster vs insulation"]
    _style_blocks(ws, bvi_layout, ws.max_column)

    # Uniform professional font across every sheet, preserving the
    # bold/italic weights set above.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(name="Arial", size=10,
                                      bold=cell.font.bold, italic=cell.font.italic)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH.resolve()}")


if __name__ == "__main__":
    main()
